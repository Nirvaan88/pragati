from flask import Blueprint, request, redirect, url_for, flash, session, make_response, send_file
from models import employees_collection
from utils import normalize_family
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
import re

export_bp = Blueprint('export', __name__)

def format_date_ddmmyyyy(date_str):
    from datetime import datetime
    if not date_str:
        return ''
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%Y/%m/%d', '%d/%m/%Y', '%Y.%m.%d', '%d.%m.%Y'):
        try:
            d = datetime.strptime(date_str, fmt)
            return d.strftime('%d-%m-%Y')
        except Exception:
            continue
    return date_str

@export_bp.route('/export_handler', methods=['POST'])
def export_handler():
    if session.get('role') != 'admin':
        flash('Unauthorized access.', 'danger')
        return redirect(url_for('auth.dashboard'))

    export_type = request.form.get('export_type')
    search = request.form.get('search', '').strip()
    selected_ids = request.form.get('selected_ids', '').split(',')

    query = {'role': {'$ne': 'admin'}}

    if selected_ids and selected_ids != ['']:
        from bson.objectid import ObjectId
        object_ids = []
        for eid in selected_ids:
            if eid:
                try:
                    object_ids.append(ObjectId(eid))
                except:
                    pass
        query['_id'] = {'$in': object_ids}
    elif search:
        exact_fields = ['employee_id', 'phone', 'gender']
        partial_fields = ['name', 'designation', 'email']
        pattern = f'\b{re.escape(search)}\b'
        query['$or'] = [
            {field: {'$regex': pattern, '$options': 'i'}} for field in exact_fields
        ] + [
            {field: {'$regex': re.escape(search), '$options': 'i'}} for field in partial_fields
        ]

    employees = list(employees_collection.find(query))

    if export_type == 'csv':
        return generate_csv(employees)
    elif export_type == 'excel':
        return generate_excel(employees)
    else:
        flash("Invalid export type", "danger")
        return redirect(url_for('admin.admin_dashboard'))

def get_formatted_relationship(relationship, gender):
    relationship = relationship.strip().lower()
    gender = gender.strip().lower() if gender else ''
    
    if relationship == 'spouse':
        return 'Wife' if gender == 'female' else 'Husband'
    elif relationship == 'child':
        return 'Daughter' if gender == 'female' else 'Son'
    return relationship.capitalize()

def get_display_age(relationship, age_str):
    if not age_str:
        return ''
    
    relationship = relationship.strip().lower()
    age_str = age_str.strip().lower()

    if relationship == 'spouse':
        return re.sub(r'\s*(years?|yrs?)$', '', age_str).strip()  # remove "years", "yr", etc.

    if relationship == 'child':
        if re.match(r'^\d+\s*(years?|yrs?)$', age_str):
            return re.sub(r'\s*(years?|yrs?)$', '', age_str).strip()
        return age_str  # e.g., "5 months", "newborn", etc.

    return age_str  # For others, return as-is

def generate_csv(employees):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'Sr. No', 'Employee Code', 'Name of Employee/Dependent', 'DOB', 'Age', 'Relation', 'Gender',
        'Designation', 'Contact No.', 'Date of Joining',
        'Sum Insured - GMC', 'Sum Insured - GPA', 'Sum Insured - GTL', 'Email ID', 'Marital Status'
    ])

    sr_no = 1
    for emp in employees:
        emp = normalize_family(emp)
        writer.writerow([
            sr_no,
            emp.get('employee_id', ''),
            emp.get('name', ''),
            format_date_ddmmyyyy(emp.get('dob', '')),
            emp.get('age', ''),
            'Employee',
            emp.get('gender', ''),
            emp.get('designation', ''),
            emp.get('phone', ''),
            format_date_ddmmyyyy(emp.get('date_of_joining', '')),
            emp.get('sum_insured_gmc', ''),
            emp.get('sum_insured_gpa', ''),
            emp.get('sum_insured_gtl', ''),
            emp.get('email', ''),
            emp.get('marital_status', '')
        ])
        if emp.get('family_members'):
            for member in emp.get('family_members', []):
                writer.writerow([
                    '', '', member.get('name', ''),
                    format_date_ddmmyyyy(member.get('date_of_birth', '')),
                    get_display_age(member.get('relationship', ''), member.get('age', '')),
                    get_formatted_relationship(member.get('relationship', ''), member.get('gender', '')),
                    member.get('gender', ''), '', '', '', '', '', '', '', ''
                ])
        writer.writerow([''] * 15)
        writer.writerow([''] * 15)
        sr_no += 1

    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=employees_nested.csv'
    response.headers['Content-type'] = 'text/csv'
    return response

def generate_excel(employees):
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Data"

    headers = [
        'Sr. No', 'Employee Code', 'Name of Employee/Dependent', 'DOB', 'Age', 'Relation', 'Gender',
        'Designation', 'Contact No.', 'Date of Joining',
        'Sum Insured - GMC', 'Sum Insured - GPA', 'Sum Insured - GTL', 'Email ID', 'Marital Status'
    ]

    header_fill = PatternFill(start_color='042351', end_color='042351', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    align_center = Alignment(horizontal='center', vertical='center')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    sr_no = 1
    row = 2
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    fill_gray = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    for emp in employees:
        emp = normalize_family(emp)
        marital_status = emp.get('marital_status', '').strip().lower()
        fill = fill_white if sr_no % 2 != 0 else fill_gray

        ws.append([
            sr_no,
            emp.get('employee_id', ''),
            emp.get('name', ''),
            format_date_ddmmyyyy(emp.get('dob', '')),
            emp.get('age', ''),
            'Employee',
            emp.get('gender', ''),
            emp.get('designation', ''),
            emp.get('phone', ''),
            format_date_ddmmyyyy(emp.get('date_of_joining', '')),
            emp.get('sum_insured_gmc', ''),
            emp.get('sum_insured_gpa', ''),
            emp.get('sum_insured_gtl', ''),
            emp.get('email', ''),
            emp.get('marital_status', '')
        ])
        for col in range(1, 16):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = thin_border
        row += 1

        if emp.get('family_members'):
            for member in emp.get('family_members', []):
                ws.append([
                    '', '', member.get('name', ''),
                    format_date_ddmmyyyy(member.get('date_of_birth', '')),
                    get_display_age(member.get('relationship', ''), member.get('age', '')),
                    get_formatted_relationship(member.get('relationship', ''), member.get('gender', '')),
                    member.get('gender', ''), '', '', '', '', '', '', '', ''
                ])
                ws[f'C{row}'].alignment = Alignment(indent=1)
                for col in range(1, 16):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = fill
                    cell.border = thin_border
                row += 1

        for _ in range(2):
            ws.append([''] * 15)
            row += 1
        sr_no += 1

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name='employees_nested.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
